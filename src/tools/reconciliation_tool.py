"""
分录对账工具模块
提供JE文件加载、TB文件加载和对账分析功能
支持大文件分批处理
"""
import pandas as pd
import numpy as np
import json
import re
import os
import tempfile
from typing import List, Dict, Optional, Tuple, Any, Iterator, Generator
from langchain.tools import tool
from coze_coding_utils.log.write_log import request_context
from coze_coding_utils.runtime_ctx.context import new_context

# 默认配置
DEFAULT_THRESHOLD = 0.01
# 默认批次大小（行数）
DEFAULT_BATCH_SIZE = 10000
# 工具返回给 LLM 的明细条数上限。完整明细写入临时文件，避免模型上下文超限。
DEFAULT_RESULT_PREVIEW_LIMIT = 5
# 内存警告阈值（MB）
MEMORY_WARNING_THRESHOLD = 500


JE_COLUMNS = {
    "book": ["账套", "公司", "工厂"],
    "voucher": ["凭证号", "凭证编号", "凭证", "记账凭证号"],
    "year": ["年", "年度"],
    "month": ["月", "月份"],
    "period": ["期间", "会计期间", "期间代码"],
    "subject": ["科目", "会计科目", "科目编码", "科目代码"],
    "account_code": ["科目编码", "总账科目"],
    "account_name": ["科目名称", "科目全称"],
    "debit": ["借方本位币", "借方本位币金额", "借方金额(本位币)", "求和项:借贷方金额(本位币)", "借方金额", "借贷方本位币"],
    "credit": ["贷方本位币", "贷方本位币金额", "贷方金额(本位币)", "求和项:借贷方金额(本位币)", "贷方金额"],
    "description": ["摘要", "说明", "描述"],
    "reversal": ["红字", "红冲", "冲销", "反方向"],
}

TB_COLUMNS = {
    "book": ["核算账套名称", "主体账套", "账套", "公司"],
    "account_code": ["科目编码", "总账科目", "科目"],
    "account_name": ["科目名称", "科目全称", "名称"],
    "subject": ["科目", "会计科目", "科目编码", "科目代码"],
    "debit": ["本期借方.1", "本期借方发生.1", "本期借方", "借方累计.1", "借方累计", "本期借方发生_1", "借贷方本位币"],
    "credit": ["本期贷方.1", "本期贷方发生.1", "本期贷方", "贷方累计.1", "贷方累计", "本期贷方发生_1", "贷方本位币"],
}


def parse_currency_value(x: Any) -> float:
    """解析货币金额值"""
    try:
        if pd.isna(x):
            return 0.0
        s = str(x).strip()
        if s == '':
            return 0.0
        s = s.replace(',', '')
        neg = False
        if s.startswith('(') and s.endswith(')'):
            neg = True
            s = s[1:-1]
        v = float(s)
        return -v if neg else v
    except Exception:
        try:
            return float(x)
        except Exception:
            return 0.0


def extract_account_code(s: Any) -> str:
    """提取科目代码"""
    s = str(s)
    m = re.search(r'(\d+(?:\.\d+)*)', s)
    if m:
        return m.group(1)
    digits = re.findall(r'\d+', s)
    return ''.join(digits)


def _is_reversal_flag(v: Any) -> bool:
    """判断是否为红字标志"""
    if v is None:
        return False
    if isinstance(v, (bool, np.bool_)):
        return bool(v)
    try:
        if isinstance(v, (int, float, np.integer, np.floating)):
            return float(v) != 0.0
    except Exception:
        pass
    s = str(v).strip()
    if s == '' or s.lower() in {'false', '0', 'no', 'n', '否', 'f'}:
        return False
    return s.lower() in {'true', '1', 'x', 'y', 'yes', '是', 't'}


def _build_header_to_idx(headers: List[str]) -> Dict[str, int]:
    """构建列名到索引的映射"""
    header_to_idx = {}
    for i, h in enumerate(headers):
        hh = str(h).strip() if h is not None else ''
        if hh and hh not in header_to_idx:
            header_to_idx[hh] = i
    return header_to_idx


def _first_existing(header_to_idx: Dict[str, int], candidates: List[str]) -> Optional[str]:
    """从候选列名中找到第一个存在于表头的列。"""
    for c in candidates:
        if c and c in header_to_idx:
            return c
    return None


def _row_has_value(row_dict: Dict[str, Any]) -> bool:
    """过滤完全空白的数据行。"""
    for value in row_dict.values():
        if value is not None and str(value).strip() != "":
            return True
    return False


def _map_row_to_standard_columns(row: Tuple[Any, ...], header_to_idx: Dict[str, int], file_type: str) -> Dict[str, Any]:
    """将一行原始 Excel 数据映射成工具内部使用的标准中文列名。"""
    columns = TB_COLUMNS if file_type == 'tb' else JE_COLUMNS
    row_dict: Dict[str, Any] = {}

    def get_value(col_name: Optional[str]) -> Any:
        if not col_name or col_name not in header_to_idx:
            return None
        idx = header_to_idx[col_name]
        return row[idx] if idx < len(row) else None

    book_col = _first_existing(header_to_idx, columns.get("book", []))
    voucher_col = _first_existing(header_to_idx, columns.get("voucher", []))
    year_col = _first_existing(header_to_idx, columns.get("year", []))
    month_col = _first_existing(header_to_idx, columns.get("month", []))
    period_col = _first_existing(header_to_idx, columns.get("period", []))
    subject_col = _first_existing(header_to_idx, columns.get("subject", []))
    account_code_col = _first_existing(header_to_idx, columns.get("account_code", []))
    account_name_col = _first_existing(header_to_idx, columns.get("account_name", []))
    debit_col = _first_existing(header_to_idx, columns.get("debit", []))
    credit_col = _first_existing(header_to_idx, columns.get("credit", []))
    description_col = _first_existing(header_to_idx, columns.get("description", []))
    reversal_col = _first_existing(header_to_idx, columns.get("reversal", []))

    if book_col:
        row_dict['账套'] = get_value(book_col)
    if voucher_col:
        row_dict['凭证号'] = get_value(voucher_col)
    if year_col:
        row_dict['年'] = get_value(year_col)
    if month_col:
        row_dict['月'] = get_value(month_col)
    if period_col:
        row_dict['期间'] = get_value(period_col)
    if subject_col:
        row_dict['科目'] = get_value(subject_col)
    if account_code_col:
        row_dict['科目编码'] = get_value(account_code_col)
    if account_name_col:
        row_dict['科目名称'] = get_value(account_name_col)
    if debit_col:
        row_dict['借方金额'] = get_value(debit_col)
    if credit_col:
        row_dict['贷方金额'] = get_value(credit_col)
    if description_col:
        row_dict['摘要'] = get_value(description_col)
    if reversal_col:
        row_dict['红字'] = get_value(reversal_col)

    return row_dict if _row_has_value(row_dict) else {}


def _read_csv_in_chunks(file_path: str, chunk_size: Optional[int] = None) -> Iterator[pd.DataFrame]:
    """读取 CSV，优先 utf-8，失败后回退到 gbk。"""
    read_kwargs = {"on_bad_lines": "skip"}
    if chunk_size:
        read_kwargs["chunksize"] = chunk_size

    last_error: Optional[Exception] = None
    for encoding in ("utf-8", "gbk"):
        try:
            chunks = pd.read_csv(file_path, encoding=encoding, **read_kwargs)
            if isinstance(chunks, pd.DataFrame):
                chunks.columns = [str(c).strip() for c in chunks.columns]
                yield chunks
            else:
                for chunk in chunks:
                    chunk.columns = [str(c).strip() for c in chunk.columns]
                    yield chunk
            return
        except UnicodeDecodeError as e:
            last_error = e
            continue
    if last_error:
        raise last_error


def _write_records_csv(prefix: str, records: List[Dict[str, Any]]) -> Optional[str]:
    """将完整明细写到唯一临时 CSV，返回文件路径；无记录则返回 None。"""
    if not records:
        return None
    path = tempfile.NamedTemporaryFile(prefix=prefix, suffix='.csv', delete=False).name
    pd.DataFrame(records).to_csv(path, index=False, encoding='utf-8')
    return path


def _preview_records(records: List[Dict[str, Any]], limit: int = DEFAULT_RESULT_PREVIEW_LIMIT) -> List[Dict[str, Any]]:
    """返回少量样例给 LLM，避免工具消息过大。"""
    return records[:limit]


def _select_excel_sheet_and_header_row(file_path: str, prefer_sheet_keywords: Optional[List[str]] = None) -> Tuple[Optional[str], int]:
    """智能选择Excel工作表和表头行"""
    prefer_sheet_keywords = prefer_sheet_keywords or []
    try:
        from openpyxl import load_workbook
        wb = load_workbook(file_path, read_only=True, data_only=True)
        best = None
        for name in wb.sheetnames:
            ws = wb[name]
            prefer = 0
            for kw in prefer_sheet_keywords:
                if kw and kw in str(name):
                    prefer += 10
            for r in range(1, 11):
                try:
                    header_row = next(ws.iter_rows(min_row=r, max_row=r, values_only=True))
                except StopIteration:
                    continue
                headers = [str(v).strip() if v is not None else '' for v in header_row]
                header_set = set(h for h in headers if h)
                score = 0
                if any(x in header_set for x in ['账套', '公司', '工厂', '核算账套名称', '主体账套']):
                    score += 5
                if any(x in header_set for x in ['科目', '会计科目', '科目编码', '会计科目末级编码', '行标签']):
                    score += 8
                if any(x in header_set for x in ['借贷方本位币', '借贷方本位币金额', '借方金额(本位币)', '求和项:借贷方金额(本位币)', '借方金额']):
                    score += 6
                if any(x in header_set for x in ['贷方本位币', '贷方本位币金额', '贷方金额(本位币)', '求和项:贷方金额(本位币)', '贷方金额']):
                    score += 6
                if any(x in header_set for x in ['期间', '月', '月份', '会计期间']):
                    score += 2
                if any(x in header_set for x in ['凭证号', '凭证编号', '凭证', '记账凭证号']):
                    score += 2
                key = (score + prefer, score, prefer, name, r)
                if best is None or key > best:
                    best = key
        wb.close()
        if best is None:
            return None, 1
        return best[3], int(best[4])
    except Exception:
        return None, 1


def _load_excel_file(file_path: str, file_type: str = 'je', target_patterns: Optional[List[str]] = None) -> pd.DataFrame:
    """加载完整文件。

    注意：对账主流程应优先使用 ``_load_excel_in_chunks``。该函数保留给 TB 小表、
    兼容旧调用和单独加载工具使用。
    """
    chunks = list(_load_excel_in_chunks(file_path, file_type=file_type, chunk_size=DEFAULT_BATCH_SIZE))
    if not chunks:
        return pd.DataFrame()
    return pd.concat(chunks, ignore_index=True)


def _load_excel_in_chunks(file_path: str, file_type: str = 'je', chunk_size: int = DEFAULT_BATCH_SIZE) -> Generator[pd.DataFrame, None, None]:
    """
    分批加载Excel文件，避免内存溢出
    
    Args:
        file_path: 文件路径
        file_type: 文件类型 ('je' 或 'tb')
        chunk_size: 每批次行数
    
    Yields:
        DataFrame批次
    """
    ext = os.path.splitext(str(file_path))[1].lower()
    
    if ext in {'.csv'}:
        # CSV 文件使用 pandas 原生 chunksize，避免全量读入内存。
        for chunk in _read_csv_in_chunks(file_path, chunk_size=chunk_size):
            yield chunk
    elif ext in {'.xlsx', '.xlsm'}:
        # Excel 使用 openpyxl read_only + iter_rows 流式读取。旧实现会先把整个
        # 工作表读入 list/DataFrame，再切 chunk；大文件会明显放大内存占用。
        from openpyxl import load_workbook

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            sheet_name, header_row_idx = _select_excel_sheet_and_header_row(
                file_path,
                prefer_sheet_keywords=['凭证', '序时', '分录'] if file_type == 'je' else ['余额', '科目']
            )
            ws = wb[sheet_name] if sheet_name in wb.sheetnames else wb.active
            row_iter = ws.iter_rows(min_row=header_row_idx, values_only=True)

            try:
                header_row = next(row_iter)
            except StopIteration:
                return

            headers = [str(v).strip() if v is not None else '' for v in header_row]
            header_to_idx = _build_header_to_idx(headers)
            data_rows: List[Dict[str, Any]] = []

            for row in row_iter:
                row_dict = _map_row_to_standard_columns(row, header_to_idx, file_type)
                if not row_dict:
                    continue
                data_rows.append(row_dict)

                if len(data_rows) >= chunk_size:
                    yield pd.DataFrame(data_rows)
                    data_rows = []

            if data_rows:
                yield pd.DataFrame(data_rows)
        finally:
            wb.close()
    else:
        return


def _filter_tb_to_leaf_accounts(tb_df: pd.DataFrame, code_col: str = '科目编码', book_col: str = '账套') -> pd.DataFrame:
    """过滤TB数据，只保留末级科目"""
    if tb_df is None or len(tb_df) == 0:
        return tb_df
    if code_col not in tb_df.columns:
        return tb_df
    
    work = tb_df.copy()
    work[code_col] = work[code_col].astype(str).str.strip()
    if book_col in work.columns:
        work[book_col] = work[book_col].astype(str).str.strip()
    else:
        work[book_col] = ''
    
    removed_total = 0
    keep_parts = []
    
    for book, g in work.groupby(book_col, dropna=False):
        codes = g[code_col].astype(str).str.strip()
        uniq = sorted(set(c for c in codes.tolist() if c and c != 'nan'))
        non_leaf = set()
        
        for i in range(len(uniq) - 1):
            cur = uniq[i]
            nxt = uniq[i + 1]
            if nxt.startswith(cur) and len(nxt) > len(cur):
                non_leaf.add(cur)
        
        if non_leaf:
            keep = ~g[code_col].isin(non_leaf)
            removed_total += int((~keep).sum())
            keep_parts.append(g[keep].copy())
        else:
            keep_parts.append(g.copy())
    
    out = pd.concat(keep_parts, ignore_index=True) if keep_parts else work
    
    if removed_total > 0:
        print(f"已按末级科目过滤TB父级科目: 移除 {removed_total} 条")
    
    return out


@tool
def load_je_data(je_file_paths: str) -> str:
    """
    加载分录(JE)数据文件（支持大文件分批处理）。
    
    Args:
        je_file_paths: JE文件路径，支持多个文件用逗号分隔
                     例如: "/path/to/je1.xlsx,/path/to/je2.xlsx"
    
    Returns:
        返回加载结果的JSON字符串，包含成功加载的文件数和数据行数
    """
    ctx = request_context.get() or new_context(method="load_je_data")
    
    try:
        file_list = [f.strip() for f in je_file_paths.split(',')]
        
        total_rows = 0
        files_loaded = 0
        columns: List[str] = []
        wrote_header = False
        temp_file = tempfile.NamedTemporaryFile(prefix='je_loaded_', suffix='.csv', delete=False).name
        
        for file_path in file_list:
            if os.path.basename(str(file_path)).startswith('~$'):
                continue
            if os.path.exists(file_path):
                file_rows = 0
                for batch in _load_excel_in_chunks(file_path, file_type='je'):
                    if batch is None or len(batch) == 0:
                        continue

                    # 解析金额
                    if '借方金额' in batch.columns:
                        batch['借方金额'] = batch['借方金额'].apply(parse_currency_value)
                    if '贷方金额' in batch.columns:
                        batch['贷方金额'] = batch['贷方金额'].apply(parse_currency_value)

                    # 提取科目代码
                    if '科目' in batch.columns:
                        batch['科目编码'] = batch['科目'].apply(extract_account_code)

                    if not columns:
                        columns = list(batch.columns)
                    batch.to_csv(temp_file, mode='a', index=False, header=not wrote_header, encoding='utf-8')
                    wrote_header = True
                    file_rows += len(batch)
                    total_rows += len(batch)

                if file_rows > 0:
                    files_loaded += 1
                    print(f"成功加载JE文件: {file_path}, 行数: {file_rows}")
            else:
                print(f"警告: JE文件不存在: {file_path}")
        
        if files_loaded == 0:
            try:
                os.unlink(temp_file)
            except OSError:
                pass
            return json.dumps({
                "success": False,
                "error": "没有成功加载任何JE文件",
                "files_loaded": 0,
                "total_rows": 0
            }, ensure_ascii=False)
        
        return json.dumps({
            "success": True,
            "message": f"成功加载 {files_loaded} 个JE文件",
            "files_loaded": files_loaded,
            "total_rows": total_rows,
            "columns": columns,
            "data_file": temp_file
        }, ensure_ascii=False)
        
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": f"加载JE文件失败: {str(e)}"
        }, ensure_ascii=False)


@tool
def load_tb_data(tb_file_path: str) -> str:
    """
    加载科目余额表(TB)数据文件（支持大文件分批处理）。
    
    Args:
        tb_file_path: TB文件路径，例如: "/path/to/tb.xlsx"
    
    Returns:
        返回加载结果的JSON字符串，包含数据行数和列信息
    """
    ctx = request_context.get() or new_context(method="load_tb_data")
    
    try:
        if not os.path.exists(tb_file_path):
            return json.dumps({
                "success": False,
                "error": f"TB文件不存在: {tb_file_path}"
            }, ensure_ascii=False)
        
        df = _load_excel_file(tb_file_path, file_type='tb')
        
        if df is None or len(df) == 0:
            return json.dumps({
                "success": False,
                "error": "TB文件为空或无法解析"
            }, ensure_ascii=False)
        
        # 解析金额
        if '借方金额' in df.columns:
            df['借方金额'] = df['借方金额'].apply(parse_currency_value)
        if '贷方金额' in df.columns:
            df['贷方金额'] = df['贷方金额'].apply(parse_currency_value)
        
        # 提取科目代码
        if '科目' in df.columns:
            df['科目编码'] = df['科目'].apply(extract_account_code)
        
        # 过滤到末级科目
        df = _filter_tb_to_leaf_accounts(df)
        
        # 保存到唯一临时文件，避免并发请求互相覆盖；使用 CSV 避免 pickle 风险。
        temp_file = tempfile.NamedTemporaryFile(prefix='tb_loaded_', suffix='.csv', delete=False).name
        df.to_csv(temp_file, index=False, encoding='utf-8')
        
        return json.dumps({
            "success": True,
            "message": "成功加载TB文件",
            "total_rows": len(df),
            "columns": list(df.columns),
            "data_file": temp_file
        }, ensure_ascii=False)
        
    except Exception as e:
        return json.dumps({
            "success": False,
            "error": f"加载TB文件失败: {str(e)}"
        }, ensure_ascii=False)


def _summarize_je_batch(je_batch: pd.DataFrame, patterns: List[str]) -> Dict[Tuple, Dict]:
    """
    对JE批次数据进行汇总
    
    Returns:
        {(账套, 科目编码): {'借方': float, '贷方': float, '条数': int}}
    """
    summary = {}
    
    for _, row in je_batch.iterrows():
        book = str(row.get('账套', '默认账套')).strip()
        code = str(row.get('科目编码', '')).strip()
        
        if patterns and not any(p in code for p in patterns):
            continue
        
        key = (book, code)
        
        if key not in summary:
            summary[key] = {'借方': 0.0, '贷方': 0.0, '条数': 0, '科目名称': ''}
        
        debit = float(row.get('借方金额', 0) or 0)
        credit = float(row.get('贷方金额', 0) or 0)
        
        # 处理红字
        is_reversal = _is_reversal_flag(row.get('红字'))
        if is_reversal:
            summary[key]['贷方'] -= debit
            summary[key]['借方'] -= credit
        else:
            summary[key]['借方'] += debit
            summary[key]['贷方'] += credit
        
        summary[key]['条数'] += 1
        
        # 记录科目名称
        if not summary[key]['科目名称'] and '科目' in row:
            summary[key]['科目名称'] = str(row.get('科目', ''))
    
    return summary


def _summarize_tb_batch(tb_batch: pd.DataFrame, patterns: List[str]) -> Dict[Tuple, Dict]:
    """
    对TB批次数据进行汇总
    
    Returns:
        {(账套, 科目编码): {'借方': float, '贷方': float, '科目名称': str}}
    """
    summary = {}
    
    for _, row in tb_batch.iterrows():
        book = str(row.get('账套', '默认账套')).strip()
        code = str(row.get('科目编码', '')).strip()
        
        if patterns and not any(p in code for p in patterns):
            continue
        
        key = (book, code)
        
        if key not in summary:
            summary[key] = {'借方': 0.0, '贷方': 0.0, '科目名称': ''}
        
        debit = float(row.get('借方金额', 0) or 0)
        credit = float(row.get('贷方金额', 0) or 0)
        
        summary[key]['借方'] += debit
        summary[key]['贷方'] += credit
        
        # 记录科目名称
        if not summary[key]['科目名称']:
            if '科目' in row:
                summary[key]['科目名称'] = str(row.get('科目', ''))
            elif '科目名称' in row:
                summary[key]['科目名称'] = str(row.get('科目名称', ''))
    
    return summary


def _merge_summary(base: Dict, incremental: Dict) -> None:
    """合并两个汇总数据"""
    for key, value in incremental.items():
        if key not in base:
            base[key] = value.copy()
        else:
            base[key]['借方'] += value['借方']
            base[key]['贷方'] += value['贷方']
            base[key]['条数'] = base[key].get('条数', 0) + value.get('条数', 0)
            if not base[key].get('科目名称') and value.get('科目名称'):
                base[key]['科目名称'] = value['科目名称']


@tool
def run_reconciliation(
    je_file_paths: str,
    tb_file_path: str,
    target_patterns: str = "",
    threshold: float = 0.01,
    check_voucher_balance: bool = True,
    check_sequence: bool = False,
    batch_size: int = DEFAULT_BATCH_SIZE
) -> str:
    """
    执行分录对账分析，将JE数据与TB数据进行比对（支持大文件分批处理）。
    
    Args:
        je_file_paths: JE文件路径，多个文件用逗号分隔
        tb_file_path: TB文件路径
        target_patterns: 要对账的科目筛选模式，用逗号分隔，例如: "1001,1002,1122"
        threshold: 差异阈值，默认0.01，金额差异小于此值视为匹配
        check_voucher_balance: 是否检查凭证借贷平衡，默认True
        check_sequence: 是否检查凭证序号连续性，默认False（大数据量时禁用）
        batch_size: 批次大小，默认10000行，大文件时可调大此值
    
    Returns:
        返回对账结果的JSON字符串，包含差异明细和统计信息
    """
    ctx = request_context.get() or new_context(method="run_reconciliation")
    
    try:
        file_list = [f.strip() for f in je_file_paths.split(',')]
        patterns = [p.strip() for p in target_patterns.split(',') if p.strip()] if target_patterns else []
        
        # 分批汇总JE数据
        print(f"开始分批加载JE数据，批次大小: {batch_size}")
        je_summary: Dict[Tuple, Dict] = {}
        je_vouchers: List[Dict] = []  # 保存凭证数据用于借贷检查
        total_je_rows = 0
        batch_count = 0
        
        for file_path in file_list:
            if not os.path.exists(file_path):
                print(f"警告: JE文件不存在: {file_path}")
                continue
            
            for batch in _load_excel_in_chunks(file_path, 'je', batch_size):
                batch_count += 1
                total_je_rows += len(batch)
                
                # 解析金额
                if '借方金额' in batch.columns:
                    batch['借方金额'] = batch['借方金额'].apply(parse_currency_value)
                if '贷方金额' in batch.columns:
                    batch['贷方金额'] = batch['贷方金额'].apply(parse_currency_value)
                if '科目' in batch.columns:
                    batch['科目编码'] = batch['科目'].apply(extract_account_code)
                
                # 汇总
                batch_summary = _summarize_je_batch(batch, patterns)
                _merge_summary(je_summary, batch_summary)
                
                # 保存凭证数据用于借贷检查
                if check_voucher_balance and '凭证号' in batch.columns:
                    cols = ['账套', '凭证号', '年', '月', '借方金额', '贷方金额']
                    if '期间' in batch.columns:
                        cols.append('期间')
                    voucher_data = batch[cols].copy()
                    je_vouchers.append(voucher_data)
                
                print(f"  JE批次 {batch_count}: {len(batch)} 行，累计 {total_je_rows} 行")
        
        print(f"JE数据加载完成: {total_je_rows} 行，{len(je_summary)} 个科目")
        
        # 加载并汇总TB数据（TB通常较小，全量加载）
        print(f"开始加载TB数据...")
        tb_df = _load_excel_file(tb_file_path, file_type='tb')
        
        if tb_df is None or len(tb_df) == 0:
            return json.dumps({
                "success": False,
                "error": "TB文件为空或无法解析"
            }, ensure_ascii=False)
        
        # 解析金额
        if '借方金额' in tb_df.columns:
            tb_df['借方金额'] = tb_df['借方金额'].apply(parse_currency_value)
        if '贷方金额' in tb_df.columns:
            tb_df['贷方金额'] = tb_df['贷方金额'].apply(parse_currency_value)
        
        # 优先使用原始科目编码列，如果没有则从科目名称提取
        has_code = '科目编码' in tb_df.columns
        if has_code:
            # 检查是否有非空值
            non_empty = tb_df['科目编码'].notna() & (tb_df['科目编码'].astype(str).str.strip() != '')
            has_code = non_empty.any()
        if not has_code and '科目' in tb_df.columns:
            tb_df['科目编码'] = tb_df['科目'].apply(extract_account_code)
        
        # 过滤末级科目
        tb_df = _filter_tb_to_leaf_accounts(tb_df)
        
        # 汇总TB
        print(f"TB数据加载完成: {len(tb_df)} 行")
        tb_summary = _summarize_tb_batch(tb_df, patterns)
        
        # 执行对账
        all_codes = set(je_summary.keys()) | set(tb_summary.keys())
        
        matched_count = 0
        differences = []
        only_in_je = []
        only_in_tb = []
        
        for key in sorted(all_codes):
            book, code = key
            
            je_data = je_summary.get(key, {'借方': 0.0, '贷方': 0.0, '条数': 0, '科目名称': ''})
            tb_data = tb_summary.get(key, {'借方': 0.0, '贷方': 0.0, '科目名称': ''})
            
            je_debit = je_data['借方']
            je_credit = je_data['贷方']
            tb_debit = tb_data['借方']
            tb_credit = tb_data['贷方']
            
            debit_diff = round(je_debit - tb_debit, 2)
            credit_diff = round(je_credit - tb_credit, 2)
            
            # 获取科目名称
            item_name = je_data.get('科目名称') or tb_data.get('科目名称') or code
            
            item = {
                '账套': book,
                '科目编码': code,
                '科目名称': item_name,
                'JE借方': je_debit,
                'JE贷方': je_credit,
                'TB借方': tb_debit,
                'TB贷方': tb_credit,
                '借方差异': debit_diff,
                '贷方差异': credit_diff,
                'JE条数': je_data.get('条数', 0)
            }
            
            if key not in je_summary:
                only_in_tb.append(item)
            elif key not in tb_summary:
                only_in_je.append(item)
            elif abs(debit_diff) <= threshold and abs(credit_diff) <= threshold:
                matched_count += 1
            else:
                differences.append(item)
        
        # 凭证借贷平衡检查（限制问题数量避免响应过大）
        voucher_issues: list = []
        MAX_VOUCHER_ISSUES = 10  # 限制最多显示10个凭证问题
        voucher_issue_count = 0
        
        if check_voucher_balance and je_vouchers:
            print("执行凭证借贷平衡检查...")
            je_voucher_df = pd.concat(je_vouchers, ignore_index=True)
            
            group_cols = ['账套', '凭证号']
            if '年' in je_voucher_df.columns:
                group_cols.append('年')
            if '月' in je_voucher_df.columns:
                group_cols.append('月')
            if '期间' in je_voucher_df.columns:
                group_cols.append('期间')
            
            grouped = je_voucher_df.groupby(group_cols).agg({
                '借方金额': 'sum',
                '贷方金额': 'sum'
            }).reset_index()
            
            for _, row in grouped.iterrows():
                debit_sum = float(row['借方金额'])
                credit_sum = float(row['贷方金额'])
                
                if abs(debit_sum - credit_sum) > threshold:
                    voucher_issue_count += 1
                    if len(voucher_issues) >= MAX_VOUCHER_ISSUES:
                        continue

                    year_val = row.get('年')
                    month_val = row.get('月')
                    period_val = row.get('期间', '')
                    
                    try:
                        if year_val is not None and month_val is not None:
                            year_month = f"{int(float(year_val))}-{int(float(month_val))}"
                        else:
                            year_month = str(period_val) if period_val else ''
                    except Exception:
                        year_month = str(period_val) if period_val else ''
                    
                    voucher_issues.append({
                        '账套': str(row['账套']) if row['账套'] is not None else '',
                        '凭证号': str(row['凭证号']) if row['凭证号'] is not None else '',
                        '期间': year_month,
                        '问题类型': '借贷不平衡',
                        '借方合计': round(debit_sum, 2),
                        '贷方合计': round(credit_sum, 2),
                        '差异': round(debit_sum - credit_sum, 2)
                    })
            
            # 计算实际检查的凭证数（不计入已跳过的）
            total_vouchers_checked = len(grouped)
            print(f"  检查 {total_vouchers_checked} 张凭证，发现 {voucher_issue_count} 个借贷不平衡问题（最多返回{MAX_VOUCHER_ISSUES}个样例）")

        result_files = {
            "differences_csv": _write_records_csv('recon_differences_', differences),
            "only_in_je_csv": _write_records_csv('recon_only_in_je_', only_in_je),
            "only_in_tb_csv": _write_records_csv('recon_only_in_tb_', only_in_tb),
            "voucher_issues_preview_csv": _write_records_csv('recon_voucher_issues_preview_', voucher_issues),
        }
        result_files = {k: v for k, v in result_files.items() if v}
        
        # 构建紧凑结果：工具消息会进入 LLM 上下文，不能返回大量明细。
        # 完整明细写入 result_files 指向的 CSV 文件。
        result = {
            "success": True,
            "message": "对账完成。为避免模型上下文超限，响应仅包含摘要和少量样例；完整明细请读取 result_files 中的 CSV 文件。",
            "summary": {
                "总科目数": len(all_codes),
                "匹配数": matched_count,
                "差异数": len(differences),
                "仅在JE中": len(only_in_je),
                "仅在TB中": len(only_in_tb),
                "凭证异常数": voucher_issue_count,
                "JE总行数": total_je_rows,
                "TB总行数": len(tb_df),
                "处理批次": batch_count,
                "返回样例条数上限": DEFAULT_RESULT_PREVIEW_LIMIT
            },
            "result_files": result_files,
            "preview": {
                "differences": _preview_records(differences),
                "only_in_je": _preview_records(only_in_je),
                "only_in_tb": _preview_records(only_in_tb),
                "voucher_issues": _preview_records(voucher_issues),
            },
            "matched_count": matched_count
        }
        
        return json.dumps(result, ensure_ascii=False, indent=2)
        
    except Exception as e:
        import traceback
        traceback.print_exc()
        return json.dumps({
            "success": False,
            "error": f"对账分析失败: {str(e)}"
        }, ensure_ascii=False)
